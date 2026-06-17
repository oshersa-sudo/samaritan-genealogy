# -*- coding: utf-8 -*-
import openpyxl, io, json, re, unicodedata
from collections import Counter, OrderedDict

FN = r"C:\Users\osher\Downloads\מיפקד.xlsx"
wb = openpyxl.load_workbook(FN, data_only=True)

NIQQUD = ''.join(chr(c) for c in range(0x0591,0x05C8))
def strip_niqqud(s):
    return ''.join(ch for ch in s if ch not in NIQQUD)
def clean(v):
    if v is None: return ''
    s = str(v).strip()
    if s.endswith('.0'):
        try:
            if float(v)==int(float(v)): s=str(int(float(v)))
        except: pass
    return strip_niqqud(s)
def norm(s):
    if not s: return ''
    s = strip_niqqud(str(s))
    s = re.sub(r'\(.*?\)', ' ', s)          # drop parenthetical alt-names for matching
    s = re.sub(r'[׳״"\'`\.,/\\-]', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s

HEADER_LABELS = {
    'השם':'name','שם':'name',
    'משפחה':'family',
    'מין':'sex',
    'תאריך לידה':'bdate',
    'מכום לידה':'bplace','מקום לידה':'bplace',
    'שם האב':'father',
    'עיסוקו':'job',
    'האם':'mother',
    'הערות':'notes',
    'שנה':'dyear','יום':'dday','חיו':'age',
    'מגורים':'residence','מספר':'num',
}

def fix_md(s):
    # canonical = MM.DD ; returns (month, day, status)
    s = clean(s)
    if not s or '.' not in s: return (None,None,'empty' if not s else 'nodot')
    try:
        a,b = s.split('.')[:2]; a=int(a); b=int(b)
    except: return (None,None,'bad')
    if a==0 or b==0: return (None,None,'incomplete')
    if a>12 and 1<=b<=12: return (b,a,'swapped_fixed')   # DD.MM -> fix
    if 1<=a<=12 and b>12:
        return (a,b,'ok') if b<=31 else (None,None,'invalid_day')
    if 1<=a<=12 and 1<=b<=12: return (a,b,'ambiguous_md')
    return (None,None,'invalid')

people = []
sheet_counts = {}
for ws in wb.worksheets:
    # find header row (contains 'השם')
    hdr_row = None; cols = {}
    for r in range(1, 8):
        vals = [clean(c) for c in next(ws.iter_rows(min_row=r,max_row=r,values_only=True))]
        if any(v=='השם' for v in vals):
            hdr_row = r
            seen_family=False; seen_mother=False
            for ci,v in enumerate(vals):
                lab = HEADER_LABELS.get(v)
                if not lab: continue
                if lab=='family':
                    cols['mfamily' if seen_mother else 'family'] = ci; seen_family=True
                elif lab=='mother':
                    cols['mother']=ci; seen_mother=True
                else:
                    cols.setdefault(lab, ci)
            # birth year = column right after bdate
            if 'bdate' in cols: cols['byear']=cols['bdate']+1
            break
    if hdr_row is None: 
        sheet_counts[ws.title]=0; continue
    cnt=0
    for row in ws.iter_rows(min_row=hdr_row+1, values_only=True):
        rv = [clean(c) for c in row]
        def g(key):
            i = cols.get(key); 
            return rv[i] if (i is not None and i < len(rv)) else ''
        name = g('name')
        if not name: continue
        m,d,st = fix_md(g('bdate'))
        rec = OrderedDict()
        rec['name']=name; rec['family']=g('family'); rec['sex']=g('sex')
        rec['byear']=g('byear'); rec['bmonth']=m; rec['bday']=d; rec['bdate_status']=st
        rec['bplace']=g('bplace'); rec['father']=g('father'); rec['job']=g('job')
        rec['mother']=g('mother'); rec['mfamily']=g('mfamily')
        rec['notes']=g('notes'); rec['dyear']=g('dyear'); rec['age']=g('age')
        rec['_sheet']=ws.title
        people.append(rec); cnt+=1
    sheet_counts[ws.title]=cnt

# dedupe: key = norm(name)|norm(father)|norm(family)|byear
dedup = OrderedDict()
for p in people:
    key = '|'.join([norm(p['name']), norm(p['father']), norm(p['family']), p['byear']])
    if key not in dedup:
        q = OrderedDict(p); q['_sheets']=[p['_sheet']]; q.pop('_sheet'); dedup[key]=q
    else:
        q = dedup[key]
        q['_sheets'].append(p['_sheet'])
        for f in ['sex','bmonth','bday','bplace','father','job','mother','mfamily','notes','dyear','age']:
            if not q.get(f) and p.get(f): q[f]=p[f]

uniq = list(dedup.values())
io.open('modern_people.json','w',encoding='utf-8').write(json.dumps(uniq, ensure_ascii=False, indent=1))

# summary
out = io.open('parse_summary.txt','w',encoding='utf-8')
out.write("rows per sheet: %s\n" % json.dumps(sheet_counts, ensure_ascii=False))
out.write("total rows: %d ; unique people: %d\n" % (len(people), len(uniq)))
fam = Counter(norm(p['family']) for p in uniq)
out.write("families (normalized): %s\n" % json.dumps(dict(fam.most_common()), ensure_ascii=False))
dst = Counter(p['bdate_status'] for p in uniq)
out.write("birthdate status: %s\n" % json.dumps(dict(dst), ensure_ascii=False))
yrs = [int(p['byear']) for p in uniq if re.match(r'^\d{4}$', p['byear'] or '')]
out.write("birth-year range: %s..%s ; people with year: %d\n" % (min(yrs) if yrs else '-', max(yrs) if yrs else '-', len(yrs)))
# how many born <= 1910 (overlap candidates)
ovl = [p for p in uniq if re.match(r'^\d{4}$', p['byear'] or '') and int(p['byear'])<=1910]
out.write("born <=1910 (overlap candidates): %d\n" % len(ovl))
out.close()
print("ok uniq", len(uniq))
